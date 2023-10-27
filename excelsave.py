import xlsxwriter
import openpyxl
from dotenv import load_dotenv
from os import getenv

load_dotenv()
DIRECTORY = getenv("DIRECTORY")


def write_in_file(data):
    book = xlsxwriter.Workbook(DIRECTORY + "data.xlsx")
    for item_page in data:
        page = book.add_worksheet(item_page["flavor"])

        page.set_column("A:A", 30)
        page.set_column("B:B", 30)
        page.set_column("C:C", 30)

        page.write(0, 0, "name")
        page.write(0, 1, "alt_name")
        page.write(0, 2, "brand")
        page.write(0, 3, "rating")
        page.write(0, 4, "real_rating")

        for row, item_row in enumerate(item_page["data"]):
            page.write(row + 1, 0, item_row["name"])
            page.write(row + 1, 1, item_row["alt_name"])
            page.write(row + 1, 2, item_row["brand"])
            page.write(row + 1, 3, item_row["rating"])
            page.write(row + 1, 4, item_row["real_rating"])

    book.close()


def read_file(flavor_name: str) -> list:
    book = openpyxl.load_workbook(DIRECTORY + "data.xlsx")
    pages_names = book.sheetnames
    top_list = []

    for page_name in pages_names:
        if page_name == flavor_name:
            for index, row in enumerate(book[flavor_name].values):
                if index == 0:
                    # пропускаем первую строку, так-как это заголовок
                    continue
                elif row[4] == 0 and index == 1:
                    if str(row[1]) == "None":
                        name = row[0]
                    else:
                        name = f"{row[0]} ({row[1]})"

                    line = f"{name} / {row[2]} ({row[4]:.3f})"
                    top_list.append(line)
                    break

                elif row[4] != 0:
                    if str(row[1]) == "None":
                        name = row[0]
                    else:
                        name = f"{row[0]} ({row[1]})"

                    line = f"{name} / {row[2]} ({row[4]:.3f})"
                    top_list.append(line)
                else:
                    break

    return top_list


def find_first_letter(letter: str) -> list:
    book = openpyxl.load_workbook(DIRECTORY + "data.xlsx")
    pages_names = book.sheetnames

    flavor_list = []
    for page in pages_names:
        if page[0] == letter.lower():
            flavor_list.append(page)

    return flavor_list


#  write_in_file(date_flavor)
#  read_file(flavor)
